import requests
import lxml.html
import openpyxl

url = "https://www.allscrabblewords.com/10-letter-words"


def getname(url):
    return url.split("/")[2].split(".")[1]


def parse(url):
    name = getname(url)
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    return tree.xpath("/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()")


def parsecsv():
    for i in range(2, 13):
        urlc = "https://www.allscrabblewords.com/{num}-letter-words".format(num=i)
        print(urlc)
        wordslist = parse(urlc)
        filename = "{num}.csv".format(num=i)
        with open(filename, "w") as csv_file:
            for str in wordslist:
                csv_file.write(str + "\n")


def parsexlsx():
    wb = openpyxl.Workbook()

    # delete sheets
    for sheetname in wb.sheetnames:
        wb.remove(wb[sheetname])

    for i in range(2, 13):
        sheetname = "Page {num}".format(num=i)
        wb.create_sheet(sheetname)
        sheet = wb[sheetname]
        urlc = "https://www.allscrabblewords.com/{num}-letter-words".format(num=i)
        print(urlc)
        wordslist = parse(urlc)
        for str in wordslist:
            cell = sheet.cell(row=wordslist.index(str) + 1, column=1)
            cell.value = str

    wb.save("words.xlsx")


def main():
    parsecsv()
    # parsexlsx()


if __name__ == "__main__":
    main()
